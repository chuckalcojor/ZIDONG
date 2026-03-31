from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import httpx

from app.config import settings
from app.services.supabase_service import SupabaseService


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: str) -> str:
    text = clean_text(value).lower()
    text = text.translate(
        str.maketrans(
            {
                "á": "a",
                "é": "e",
                "í": "i",
                "ó": "o",
                "ú": "u",
                "ñ": "n",
            }
        )
    )
    compact = []
    prev_space = False
    for char in text:
        if char.isalnum():
            compact.append(char)
            prev_space = False
        elif not prev_space:
            compact.append(" ")
            prev_space = True
    return "".join(compact).strip()


def placeholder_phone_from_name(name: str) -> str:
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:10]
    return f"000{digest}"


def placeholder_phone_from_code(code: str) -> str:
    digits = "".join(char for char in code if char.isdigit())
    if not digits:
        return placeholder_phone_from_name(code)
    return f"57{digits.zfill(10)[:10]}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import A3 route assignments workbook into Supabase")
    parser.add_argument("--excel", required=True, help="Absolute path to A3 VETERINARIA.xlsx")
    parser.add_argument("--sheet", default="Hoja1", help="Worksheet name")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    load_dotenv()

    workbook_path = Path(args.excel)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[args.sheet]

    supabase = SupabaseService(
        base_url=settings.supabase_url,
        service_role_key=settings.supabase_service_role_key,
    )

    courier_names: set[str] = set()
    clients_to_import: list[dict[str, str]] = []

    for row_index in range(2, ws.max_row + 1):
        code = clean_text(ws.cell(row_index, 1).value)
        clinic_name = clean_text(ws.cell(row_index, 2).value)
        tax_id = clean_text(ws.cell(row_index, 3).value)
        address = clean_text(ws.cell(row_index, 4).value)
        city = clean_text(ws.cell(row_index, 5).value)
        zone = clean_text(ws.cell(row_index, 6).value)
        courier_name = clean_text(ws.cell(row_index, 7).value)

        if not code or not clinic_name or not address:
            continue

        if courier_name:
            courier_names.add(courier_name)

        clients_to_import.append(
            {
                "external_code": code,
                "clinic_name": clinic_name,
                "tax_id": tax_id,
                "address": address,
                "city": city,
                "zone": zone,
                "courier_name": courier_name,
                "phone": placeholder_phone_from_code(code),
            }
        )

    try:
        existing_couriers = supabase.fetch_rows("couriers", {"select": "id,name,phone", "limit": "2000"})
    except httpx.HTTPStatusError as exc:
        print(
            json.dumps(
                {
                    "error": "Supabase schema for couriers is not ready",
                    "status_code": exc.response.status_code,
                },
                ensure_ascii=True,
            )
        )
        return

    courier_map = {clean_text(row.get("name")): row.get("id") for row in existing_couriers if row.get("name")}

    couriers_payload = []
    for name in sorted(courier_names):
        if name in courier_map:
            continue
        couriers_payload.append(
            {
                "name": name,
                "phone": placeholder_phone_from_name(name),
                "availability": "available",
                "is_active": True,
            }
        )

    if couriers_payload:
        supabase.insert_rows("couriers", couriers_payload)
        refreshed = supabase.fetch_rows("couriers", {"select": "id,name", "limit": "2000"})
        courier_map = {clean_text(row.get("name")): row.get("id") for row in refreshed if row.get("name")}

    clients_payload = []
    for item in clients_to_import:
        clients_payload.append(
            {
                "external_code": item["external_code"],
                "clinic_name": item["clinic_name"],
                "tax_id": item["tax_id"] or None,
                "phone": item["phone"],
                "address": item["address"],
                "city": item["city"] or None,
                "zone": item["zone"] or None,
                "billing_type": "cash",
                "is_active": True,
            }
        )

    dedup_clients = {item["phone"]: item for item in clients_payload}
    inserted_clients = supabase.insert_rows(
        "clients",
        list(dedup_clients.values()),
        upsert=True,
        on_conflict="phone",
    )

    client_id_by_phone = {
        clean_text(row.get("phone")): row.get("id")
        for row in inserted_clients
        if row.get("phone") and row.get("id")
    }

    assignment_payload = []
    for item in clients_to_import:
        client_id = client_id_by_phone.get(item["phone"])
        courier_id = courier_map.get(clean_text(item["courier_name"]))
        if not client_id or not courier_id:
            continue
        assignment_payload.append(
            {
                "client_id": client_id,
                "courier_id": courier_id,
                "assigned_by": "a3_veterinaria_excel_import",
            }
        )

    if assignment_payload:
        dedup_assignments = {item["client_id"]: item for item in assignment_payload}
        supabase.insert_rows(
            "client_courier_assignment",
            list(dedup_assignments.values()),
            upsert=True,
            on_conflict="client_id",
        )

    knowledge_rows = []
    for item in clients_to_import:
        clinic_key = normalize_key(item["clinic_name"])
        if not clinic_key:
            continue
        knowledge_rows.append(
            {
                "clinic_key": clinic_key,
                "clinic_name": item["clinic_name"],
                "is_registered": True,
                "is_new_client": False,
                "address": item["address"],
                "locality": item["city"] or None,
                "phone": None,
                "email": None,
                "payment_policy": None,
                "result_delivery_mode": None,
                "sources_json": ["a3_veterinaria_excel"],
                "source_excel": str(workbook_path),
            }
        )

    if knowledge_rows:
        dedup_knowledge = {row["clinic_key"]: row for row in knowledge_rows}
        try:
            supabase.insert_rows(
                "clients_a3_knowledge",
                list(dedup_knowledge.values()),
                upsert=True,
                on_conflict="clinic_key",
            )
        except httpx.HTTPStatusError:
            pass

    print(
        json.dumps(
            {
                "excel_path": str(workbook_path),
                "rows_processed": len(clients_to_import),
                "couriers_created": len(couriers_payload),
                "clients_upserted": len(dedup_clients),
                "assignments_upserted": len({item["client_id"] for item in assignment_payload}),
            },
            ensure_ascii=True,
        )
    )


if __name__ == "__main__":
    main()
