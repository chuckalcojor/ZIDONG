from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: Any) -> str:
    text = clean_text(value).lower()
    replacements = str.maketrans(
        {
            "a": "a",
            "e": "e",
            "i": "i",
            "o": "o",
            "u": "u",
            "n": "n",
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
            "ñ": "n",
        }
    )
    text = text.translate(replacements)
    text = re.sub(r"[^a-z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_ascii(value: Any) -> str:
    return clean_text(value).encode("ascii", "ignore").decode("ascii")


def first_non_empty(*values: Any) -> str:
    for value in values:
        if clean_text(value):
            return clean_text(value)
    return ""


def normalize_bool_option(value: Any) -> str:
    normalized = normalize_key(value)
    if not normalized:
        return ""
    yes_values = {"si", "s", "yes", "true", "1", "ok", "x", "registrado", "ingresado"}
    no_values = {"no", "n", "false", "0", "pendiente"}
    if normalized in yes_values:
        return "si"
    if normalized in no_values:
        return "no"
    return ""


def normalize_client_type_option(value: Any) -> str:
    normalized = normalize_key(value)
    if not normalized:
        return ""
    if "persona" in normalized:
        return "es_persona"
    if "empresa" in normalized:
        return "empresa"
    if "otro" in normalized:
        return "otro"
    return ""


def normalize_vat_regime_option(value: Any) -> str:
    normalized = normalize_key(value)
    if not normalized:
        return ""
    if "no" in normalized and "responsable" in normalized:
        return "no_responsable_iva"
    if "responsable" in normalized:
        return "responsable_iva"
    return ""


def normalize_timestamp_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, time.min).isoformat()
    return clean_text(value)


def normalize_date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def normalize_time_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.time().isoformat(timespec="minutes")
    if isinstance(value, time):
        return value.isoformat(timespec="minutes")
    return clean_text(value)


def get_header_map(sheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(sheet[1]):
        key = normalize_key(cell.value)
        if key:
            header_map[key] = idx
    return header_map


def get_cell(row: tuple[Any, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return clean_text(row[idx])


def row_has_data(row: tuple[Any, ...]) -> bool:
    return any(clean_text(cell) for cell in row)


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        pragma foreign_keys=on;

        create table if not exists meta (
            key text primary key,
            value text not null
        );

        create table if not exists sheet_profile (
            sheet_name text primary key,
            non_empty_rows integer not null,
            max_rows integer not null,
            max_columns integer not null
        );

        create table if not exists clinic_master (
            clinic_key text primary key,
            clinic_name text not null,
            is_registered integer not null default 0,
            is_new_client integer not null default 0,
            address text,
            locality text,
            phone text,
            email text,
            payment_policy text,
            result_delivery_mode text,
            client_code text,
            commercial_name text,
            client_type text,
            billing_email text,
            vat_regime text,
            electronic_invoicing text,
            invoicing_rut_url text,
            registration_timestamp text,
            registration_date text,
            registration_time text,
            observations text,
            entered_flag text,
            sources_json text not null
        );

        create table if not exists clinic_professional (
            id integer primary key autoincrement,
            clinic_key text not null,
            professional_name text,
            professional_card text,
            source_sheet text not null,
            foreign key (clinic_key) references clinic_master(clinic_key)
        );

        create table if not exists sample_status_event (
            id integer primary key autoincrement,
            sheet_name text not null,
            clinic_key text,
            clinic_name_raw text,
            patient_name text,
            exam_code text,
            exam_number text,
            pending_exam text,
            status_bucket text not null,
            reason text,
            registered_flag text,
            observation text
        );

        create index if not exists idx_clinic_master_name on clinic_master(clinic_name);
        create index if not exists idx_sample_status_clinic on sample_status_event(clinic_key);
        create index if not exists idx_sample_status_bucket on sample_status_event(status_bucket);
        """
    )


def upsert_clinic(
    clinics: dict[str, dict[str, Any]],
    *,
    clinic_name: str,
    source_sheet: str,
    is_registered: bool = False,
    is_new_client: bool = False,
    address: str = "",
    locality: str = "",
    phone: str = "",
    email: str = "",
    payment_policy: str = "",
    result_delivery_mode: str = "",
    client_code: str = "",
    commercial_name: str = "",
    client_type: str = "",
    billing_email: str = "",
    vat_regime: str = "",
    electronic_invoicing: str = "",
    invoicing_rut_url: str = "",
    registration_timestamp: str = "",
    registration_date: str = "",
    registration_time: str = "",
    observations: str = "",
    entered_flag: str = "",
) -> str:
    clinic_key = normalize_key(clinic_name)
    if not clinic_key:
        return ""

    base = clinics.setdefault(
        clinic_key,
        {
            "clinic_key": clinic_key,
            "clinic_name": clinic_name,
            "is_registered": False,
            "is_new_client": False,
            "address": "",
            "locality": "",
            "phone": "",
            "email": "",
            "payment_policy": "",
            "result_delivery_mode": "",
            "client_code": "",
            "commercial_name": "",
            "client_type": "",
            "billing_email": "",
            "vat_regime": "",
            "electronic_invoicing": "",
            "invoicing_rut_url": "",
            "registration_timestamp": "",
            "registration_date": "",
            "registration_time": "",
            "observations": "",
            "entered_flag": "",
            "sources": set(),
        },
    )

    if len(clinic_name) > len(base["clinic_name"]):
        base["clinic_name"] = clinic_name

    base["is_registered"] = base["is_registered"] or is_registered
    base["is_new_client"] = base["is_new_client"] or is_new_client

    base["address"] = first_non_empty(base["address"], address)
    base["locality"] = first_non_empty(base["locality"], locality)
    base["phone"] = first_non_empty(base["phone"], phone)
    base["email"] = first_non_empty(base["email"], email)
    base["payment_policy"] = first_non_empty(base["payment_policy"], payment_policy)
    base["result_delivery_mode"] = first_non_empty(base["result_delivery_mode"], result_delivery_mode)
    base["client_code"] = first_non_empty(base["client_code"], client_code)
    base["commercial_name"] = first_non_empty(base["commercial_name"], commercial_name)
    base["client_type"] = first_non_empty(base["client_type"], client_type)
    base["billing_email"] = first_non_empty(base["billing_email"], billing_email)
    base["vat_regime"] = first_non_empty(base["vat_regime"], vat_regime)
    base["electronic_invoicing"] = first_non_empty(
        base["electronic_invoicing"],
        electronic_invoicing,
    )
    base["invoicing_rut_url"] = first_non_empty(base["invoicing_rut_url"], invoicing_rut_url)
    base["registration_timestamp"] = first_non_empty(
        base["registration_timestamp"],
        registration_timestamp,
    )
    base["registration_date"] = first_non_empty(base["registration_date"], registration_date)
    base["registration_time"] = first_non_empty(base["registration_time"], registration_time)
    base["observations"] = first_non_empty(base["observations"], observations)
    base["entered_flag"] = first_non_empty(base["entered_flag"], entered_flag)
    base["sources"].add(source_sheet)

    return clinic_key


def main() -> None:
    parser = argparse.ArgumentParser(description="Build normalized index from Clientes A3 workbook")
    parser.add_argument("--excel", required=True, help="Absolute path to source xlsx")
    parser.add_argument(
        "--sqlite",
        default=str(Path(__file__).resolve().parents[1] / ".cache" / "clientes_a3_index.sqlite"),
        help="Output SQLite path",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    sqlite_path = Path(args.sqlite)
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(excel_path, data_only=True)

    conn = sqlite3.connect(sqlite_path)
    conn.row_factory = sqlite3.Row

    conn.executescript(
        """
        drop table if exists clinic_professional;
        drop table if exists sample_status_event;
        drop table if exists clinic_master;
        drop table if exists sheet_profile;
        drop table if exists meta;
        """
    )

    create_schema(conn)

    clinics: dict[str, dict[str, Any]] = {}
    professionals: list[tuple[str, str, str, str]] = []
    samples: list[tuple[str, str, str, str, str, str, str, str, str, str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        non_empty_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row_has_data(row))
        conn.execute(
            "insert into sheet_profile(sheet_name, non_empty_rows, max_rows, max_columns) values (?, ?, ?, ?)",
            (sheet_name, non_empty_rows, ws.max_row, ws.max_column),
        )

    if "Clientes" in wb.sheetnames:
        ws = wb["Clientes"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Nombre")))
            if not clinic_name:
                continue
            clinic_key = upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Clientes",
                is_registered=True,
                client_code=first_non_empty(
                    get_cell(row, hmap.get(normalize_key("C"))),
                    get_cell(row, hmap.get(normalize_key("kp"))),
                ),
                phone=get_cell(row, hmap.get(normalize_key("celular"))),
                email=get_cell(row, hmap.get(normalize_key("Correo"))),
                payment_policy=get_cell(row, hmap.get(normalize_key("Pago"))),
                result_delivery_mode=get_cell(row, hmap.get(normalize_key("Medio de envio de examenes"))),
            )
            professional_name = get_cell(row, hmap.get(normalize_key("Medico veterinario")))
            professional_card = get_cell(row, hmap.get(normalize_key("N TP")))
            if clinic_key and (professional_name or professional_card):
                professionals.append((clinic_key, professional_name, professional_card, "Clientes"))

    if "Copia de Clientes" in wb.sheetnames:
        ws = wb["Copia de Clientes"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Nombre")))
            if not clinic_name:
                continue
            clinic_key = upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Copia de Clientes",
                is_registered=True,
                client_code=first_non_empty(
                    get_cell(row, hmap.get(normalize_key("C"))),
                    get_cell(row, hmap.get(normalize_key("kp"))),
                ),
                phone=get_cell(row, hmap.get(normalize_key("celular"))),
                email=get_cell(row, hmap.get(normalize_key("Correo"))),
                payment_policy=get_cell(row, hmap.get(normalize_key("Pago"))),
                result_delivery_mode=get_cell(row, hmap.get(normalize_key("Medio de envio de examenes"))),
                observations=get_cell(row, hmap.get(normalize_key("Observaciones"))),
            )
            professional_name = get_cell(row, hmap.get(normalize_key("Medico veterinario")))
            professional_card = get_cell(row, hmap.get(normalize_key("N TP")))
            if clinic_key and (professional_name or professional_card):
                professionals.append((clinic_key, professional_name, professional_card, "Copia de Clientes"))

    if "Clientes Activos 2025" in wb.sheetnames:
        ws = wb["Clientes Activos 2025"]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = clean_text(row[2]) if len(row) > 2 else ""
            if not clinic_name:
                continue
            clinic_key = upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Clientes Activos 2025",
                is_registered=True,
                phone=clean_text(row[8]) if len(row) > 8 else "",
                email=clean_text(row[7]) if len(row) > 7 else "",
                payment_policy=clean_text(row[0]) if len(row) > 0 else "",
                result_delivery_mode=clean_text(row[3]) if len(row) > 3 else "",
            )
            professional_name = clean_text(row[5]) if len(row) > 5 else ""
            professional_card = clean_text(row[6]) if len(row) > 6 else ""
            if clinic_key and (professional_name or professional_card):
                professionals.append((clinic_key, professional_name, professional_card, "Clientes Activos 2025"))

    if "Nuevos" in wb.sheetnames:
        ws = wb["Nuevos"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Nombre de la veterinaria o medico veterinario")))
            if not clinic_name:
                continue
            clinic_key = upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Nuevos",
                is_new_client=True,
                client_code=get_cell(row, hmap.get(normalize_key("Codigo"))),
                commercial_name=get_cell(row, hmap.get(normalize_key("Nombre Comercial"))),
                client_type=normalize_client_type_option(
                    get_cell(row, hmap.get(normalize_key("Tipo")))
                ),
                address=get_cell(row, hmap.get(normalize_key("Direccion y ubicacion en Google Maps"))),
                locality=get_cell(row, hmap.get(normalize_key("Barrio y Localidad"))),
                phone=get_cell(row, hmap.get(normalize_key("N Celular"))),
                email=get_cell(row, hmap.get(normalize_key("Email"))),
                billing_email=get_cell(
                    row,
                    hmap.get(normalize_key("Correo (En el cual te llegaran las facturas)")),
                ),
                vat_regime=normalize_vat_regime_option(
                    get_cell(row, hmap.get(normalize_key("Tipo de regimen IVA")))
                ),
                invoicing_rut_url=get_cell(row, hmap.get(normalize_key("Informacion en RUT"))),
                registration_timestamp=normalize_timestamp_value(
                    row[hmap.get(normalize_key("Marca temporal"))]
                    if hmap.get(normalize_key("Marca temporal")) is not None
                    and hmap.get(normalize_key("Marca temporal")) < len(row)
                    else ""
                ),
                entered_flag=normalize_bool_option(
                    get_cell(row, hmap.get(normalize_key("Registrado")))
                ),
            )
            professional_name = get_cell(row, hmap.get(normalize_key("Medico Veterinario")))
            professional_card = get_cell(row, hmap.get(normalize_key("N Tarjeta Profesional")))
            if clinic_key and (professional_name or professional_card):
                professionals.append((clinic_key, professional_name, professional_card, "Nuevos"))

    if "Base" in wb.sheetnames:
        ws = wb["Base"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Nombre completo de la veterinaria")))
            if not clinic_name:
                continue
            clinic_key = upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Base",
                is_new_client=True,
                address=get_cell(row, hmap.get(normalize_key("Direccion, Barrio y Localidad"))),
                phone=get_cell(row, hmap.get(normalize_key("N Celular de comunicacion"))),
                email=get_cell(row, hmap.get(normalize_key("Correo o WhatsApp"))),
                billing_email=get_cell(row, hmap.get(normalize_key("Correo o WhatsApp"))),
                result_delivery_mode=get_cell(row, hmap.get(normalize_key("Medio por el cual requiere que se envio los resultados."))),
                electronic_invoicing=normalize_bool_option(
                    get_cell(row, hmap.get(normalize_key("Facturacion Electronica")))
                ),
                invoicing_rut_url=get_cell(
                    row,
                    hmap.get(normalize_key("Si deseas factura electronica adjuntar el Rut")),
                ),
                registration_timestamp=normalize_timestamp_value(
                    row[hmap.get(normalize_key("Marca temporal"))]
                    if hmap.get(normalize_key("Marca temporal")) is not None
                    and hmap.get(normalize_key("Marca temporal")) < len(row)
                    else ""
                ),
                observations=get_cell(row, hmap.get(normalize_key("Observaciones"))),
            )
            professional_name = get_cell(row, hmap.get(normalize_key("Nombre completo del medico")))
            professional_card = get_cell(row, hmap.get(normalize_key("N Tarjeta profesional")))
            if clinic_key and (professional_name or professional_card):
                professionals.append((clinic_key, professional_name, professional_card, "Base"))

    if "Registro de clientes nuevos" in wb.sheetnames:
        ws = wb["Registro de clientes nuevos"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Veterinaria o medico veterinario")))
            if not clinic_name:
                continue
            upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Registro de clientes nuevos",
                is_new_client=True,
                address=get_cell(row, hmap.get(normalize_key("Direccion"))),
                locality=get_cell(row, hmap.get(normalize_key("Barrio y localidad"))),
                phone=get_cell(row, hmap.get(normalize_key("Celular o Telefono"))),
                registration_date=normalize_date_value(
                    row[hmap.get(normalize_key("Fecha"))]
                    if hmap.get(normalize_key("Fecha")) is not None
                    and hmap.get(normalize_key("Fecha")) < len(row)
                    else ""
                ),
                registration_time=normalize_time_value(
                    row[hmap.get(normalize_key("Hora"))]
                    if hmap.get(normalize_key("Hora")) is not None
                    and hmap.get(normalize_key("Hora")) < len(row)
                    else ""
                ),
                observations=get_cell(
                    row,
                    hmap.get(normalize_key("Informacion suministrada")),
                ),
            )

    if "Medios Veterinarios" in wb.sheetnames:
        ws = wb["Medios Veterinarios"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Nombre de la veterinaria")))
            if not clinic_name:
                continue
            clinic_key = upsert_clinic(
                clinics,
                clinic_name=clinic_name,
                source_sheet="Medios Veterinarios",
                is_registered=True,
                client_code=get_cell(row, hmap.get(normalize_key("Codigo Annar"))),
                entered_flag=normalize_bool_option(
                    get_cell(row, hmap.get(normalize_key("Ingresado")))
                ),
            )
            professional_name = get_cell(row, hmap.get(normalize_key("Nombre de medico veterinario")))
            professional_card = get_cell(row, hmap.get(normalize_key("N Tarjeta profesional")))
            if clinic_key and (professional_name or professional_card):
                professionals.append((clinic_key, professional_name, professional_card, "Medios Veterinarios"))

    if "Remisiones" in wb.sheetnames:
        ws = wb["Remisiones"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Veterinaria")))
            clinic_key = normalize_key(clinic_name) if clinic_name else ""
            samples.append(
                (
                    "Remisiones",
                    clinic_key,
                    clinic_name,
                    get_cell(row, hmap.get(normalize_key("Nombre del paciente"))),
                    get_cell(row, hmap.get(normalize_key("Codigo Examen"))),
                    get_cell(row, hmap.get(normalize_key("N Examen"))),
                    "",
                    "submitted",
                    "",
                    get_cell(row, hmap.get(normalize_key("Registrado"))),
                    "",
                )
            )

    if "Muestras Pendientes o Contramue" in wb.sheetnames:
        ws = wb["Muestras Pendientes o Contramue"]
        hmap = get_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row_has_data(row):
                continue
            clinic_name = get_cell(row, hmap.get(normalize_key("Veterinaria")))
            clinic_key = normalize_key(clinic_name) if clinic_name else ""
            samples.append(
                (
                    "Muestras Pendientes o Contramue",
                    clinic_key,
                    clinic_name,
                    get_cell(row, hmap.get(normalize_key("Paciente"))),
                    "",
                    get_cell(row, hmap.get(normalize_key("Codigo Interno"))),
                    get_cell(row, hmap.get(normalize_key("Examen Pendiente"))),
                    "pending_issue",
                    get_cell(row, hmap.get(normalize_key("Motivo"))),
                    get_cell(row, hmap.get(normalize_key("Revisado"))),
                    get_cell(row, hmap.get(normalize_key("Observacion"))),
                )
            )

    for clinic in clinics.values():
        conn.execute(
            """
            insert into clinic_master(
                clinic_key,
                clinic_name,
                is_registered,
                is_new_client,
                address,
                locality,
                phone,
                email,
                payment_policy,
                result_delivery_mode,
                client_code,
                commercial_name,
                client_type,
                billing_email,
                vat_regime,
                electronic_invoicing,
                invoicing_rut_url,
                registration_timestamp,
                registration_date,
                registration_time,
                observations,
                entered_flag,
                sources_json
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                clinic["clinic_key"],
                clinic["clinic_name"],
                1 if clinic["is_registered"] else 0,
                1 if clinic["is_new_client"] else 0,
                clinic["address"],
                clinic["locality"],
                clinic["phone"],
                clinic["email"],
                clinic["payment_policy"],
                clinic["result_delivery_mode"],
                clinic["client_code"],
                clinic["commercial_name"],
                clinic["client_type"],
                clinic["billing_email"],
                clinic["vat_regime"],
                clinic["electronic_invoicing"],
                clinic["invoicing_rut_url"],
                clinic["registration_timestamp"],
                clinic["registration_date"],
                clinic["registration_time"],
                clinic["observations"],
                clinic["entered_flag"],
                json.dumps(sorted(clinic["sources"]), ensure_ascii=True),
            ),
        )

    dedup_professionals = {
        (
            clinic_key,
            normalize_key(professional_name),
            normalize_key(professional_card),
            source_sheet,
        ): (clinic_key, professional_name, professional_card, source_sheet)
        for clinic_key, professional_name, professional_card, source_sheet in professionals
        if clinic_key and (clean_text(professional_name) or clean_text(professional_card))
    }

    for clinic_key, professional_name, professional_card, source_sheet in dedup_professionals.values():
        conn.execute(
            """
            insert into clinic_professional(clinic_key, professional_name, professional_card, source_sheet)
            values (?, ?, ?, ?)
            """,
            (clinic_key, professional_name, professional_card, source_sheet),
        )

    conn.executemany(
        """
        insert into sample_status_event(
            sheet_name,
            clinic_key,
            clinic_name_raw,
            patient_name,
            exam_code,
            exam_number,
            pending_exam,
            status_bucket,
            reason,
            registered_flag,
            observation
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        samples,
    )

    conn.execute("insert into meta(key, value) values (?, ?)", ("source_excel", str(excel_path)))
    conn.execute("insert into meta(key, value) values (?, ?)", ("clinic_count", str(len(clinics))))
    conn.execute("insert into meta(key, value) values (?, ?)", ("professionals_count", str(len(dedup_professionals))))
    conn.execute("insert into meta(key, value) values (?, ?)", ("sample_events_count", str(len(samples))))
    conn.commit()

    summary = {
        "source_excel": str(excel_path),
        "sqlite_output": str(sqlite_path),
        "clinic_count": len(clinics),
        "professional_count": len(dedup_professionals),
        "sample_event_count": len(samples),
    }
    print(json.dumps(summary, ensure_ascii=True))


if __name__ == "__main__":
    main()
