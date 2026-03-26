from __future__ import annotations

import hashlib
import json
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import httpx
from dotenv import load_dotenv
from openpyxl import load_workbook

from app.config import settings
from app.services.supabase_service import SupabaseService


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def placeholder_phone_from_name(name: str) -> str:
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:10]
    return f"000{digest}"


def placeholder_phone_from_code(code: str) -> str:
    digits = "".join(char for char in str(code) if char.isdigit())
    return f"57{digits.zfill(10)[:10]}"


def main() -> None:
    load_dotenv()

    workbook_path = Path(
        r"C:\Users\Artel\Downloads\LABERIT A3 VETERINARIA\Informacion\Relacion Clientes.xlsx"
    )
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Hoja1"]

    supabase = SupabaseService(
        base_url=settings.supabase_url,
        service_role_key=settings.supabase_service_role_key,
    )

    courier_names: set[str] = set()
    client_rows: list[dict[str, object]] = []

    for row_index in range(2, ws.max_row + 1):
        code = ws.cell(row_index, 1).value
        clinic_name = normalize_text(ws.cell(row_index, 2).value)
        tax_id = normalize_text(ws.cell(row_index, 3).value)
        address = normalize_text(ws.cell(row_index, 4).value)
        city = normalize_text(ws.cell(row_index, 5).value)
        zone = normalize_text(ws.cell(row_index, 6).value)
        courier_name = normalize_text(ws.cell(row_index, 7).value)

        if not code or not clinic_name or not address:
            continue

        if courier_name and courier_name.lower() not in {"none", "no aplica"}:
            courier_names.add(courier_name)

        client_rows.append(
            {
                "external_code": str(code),
                "clinic_name": clinic_name,
                "tax_id": tax_id,
                "phone": placeholder_phone_from_code(str(code)),
                "address": address,
                "city": city,
                "zone": zone,
                "billing_type": "cash",
                "is_active": True,
                "_courier_name": courier_name,
            }
        )

    try:
        existing_couriers = supabase.fetch_rows("couriers", {"select": "id,name,phone", "limit": "2000"})
    except httpx.HTTPStatusError as exc:
        print(
            json.dumps(
                {
                    "error": "Supabase schema not ready. Run 001_v1_core_schema.sql first.",
                    "status_code": exc.response.status_code,
                },
                ensure_ascii=True,
            )
        )
        return
    courier_map = {row["name"]: row["id"] for row in existing_couriers if row.get("name")}

    to_create_couriers: list[dict[str, object]] = []
    for name in sorted(courier_names):
        if name not in courier_map:
            to_create_couriers.append(
                {
                    "name": name,
                    "phone": placeholder_phone_from_name(name),
                    "availability": "available",
                    "is_active": True,
                }
            )

    if to_create_couriers:
        supabase.insert_rows("couriers", to_create_couriers)
        refreshed = supabase.fetch_rows("couriers", {"select": "id,name", "limit": "2000"})
        courier_map = {row["name"]: row["id"] for row in refreshed if row.get("name")}

    client_payload = []
    for row in client_rows:
        payload = {
            "clinic_name": row["clinic_name"],
            "phone": row["phone"],
            "address": row["address"],
            "zone": row["zone"],
            "billing_type": row["billing_type"],
            "is_active": row["is_active"],
        }
        client_payload.append(payload)

    unique_by_phone: dict[str, dict[str, object]] = {}
    for payload in client_payload:
        unique_by_phone[str(payload["phone"])] = payload
    dedup_client_payload = list(unique_by_phone.values())

    inserted_clients = supabase.insert_rows(
        "clients",
        dedup_client_payload,
        upsert=True,
        on_conflict="phone",
    )

    client_map = {row.get("phone"): row.get("id") for row in inserted_clients}
    assignments = []
    for row in client_rows:
        phone = row["phone"]
        courier_name = row.get("_courier_name")
        client_id = client_map.get(phone)
        courier_id = courier_map.get(courier_name) if courier_name else None

        if not client_id or not courier_id:
            continue

        assignments.append(
            {
                "client_id": client_id,
                "courier_id": courier_id,
                "assigned_by": "excel_import",
            }
        )

    if assignments:
        supabase.insert_rows(
            "client_courier_assignment",
            assignments,
            upsert=True,
            on_conflict="client_id",
        )

    summary = {
        "clients_processed": len(client_rows),
        "clients_upserted": len(dedup_client_payload),
        "couriers_created": len(to_create_couriers),
        "assignments_upserted": len(assignments),
    }
    print(json.dumps(summary, ensure_ascii=True))


if __name__ == "__main__":
    main()
